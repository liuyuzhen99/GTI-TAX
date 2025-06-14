import pandas as pd
from openpyxl import load_workbook, Workbook


class ReadExcelClass:
    def __init__(self, path=''):
        self.path = path
        self.dataframe = pd.DataFrame()
        self.wb = Workbook()

    def read_template_file(self):
        self.wb = load_workbook(self.path)

    def read_GTI_excel_file(self):
        self.dataframe = pd.read_excel(self.path)
